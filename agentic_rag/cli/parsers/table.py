"""CSV/TSV/Excel parsing without pandas or macro execution."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from ..cancellation import CancellationToken
from ..errors import ConfigurationError
from ..models import ParsedDocument


MAX_ROWS = 20_000
MAX_COLUMNS = 256


def _rows_to_markdown(rows: list[list[object]], cancel: CancellationToken) -> str:
    if not rows:
        raise ConfigurationError("Table contains no values")
    width = min(max(len(row) for row in rows), MAX_COLUMNS)
    normalized = [[str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")[:2000] for value in row[:width]] + [""] * max(0, width - len(row)) for row in rows]
    header = normalized[0]
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]
    for row in normalized[1:]:
        cancel.checkpoint()
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def parse_table(path: Path, cancel: CancellationToken) -> ParsedDocument:
    cancel.checkpoint()
    if path.stat().st_size > 30 * 1024 * 1024:
        raise ConfigurationError("Table file exceeds the 30MB limit")
    if path.suffix.lower() in {".csv", ".tsv"}:
        text = path.read_text("utf-8-sig", errors="replace")
        dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
        rows = []
        for index, row in enumerate(csv.reader(io.StringIO(text), dialect=dialect)):
            cancel.checkpoint()
            if index >= MAX_ROWS:
                raise ConfigurationError(f"Table exceeds the {MAX_ROWS} row limit")
            rows.append(row)
        rendered = _rows_to_markdown(rows, cancel)
        return ParsedDocument(path.name, [{"page": 1, "text": rendered}], [], "local_delimited_table")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConfigurationError("Excel support requires the AutoMemory CLI dependencies") from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise ConfigurationError(f"Excel workbook could not be read: {type(exc).__name__}") from exc
    pages = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, 1):
            rows = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                cancel.checkpoint()
                if index >= MAX_ROWS:
                    raise ConfigurationError(f"Worksheet {sheet.title} exceeds the {MAX_ROWS} row limit")
                rows.append(list(row[:MAX_COLUMNS]))
            if rows and any(any(value not in (None, "") for value in row) for row in rows):
                pages.append({"page": sheet_index, "text": f"## Sheet: {sheet.title}\n\n{_rows_to_markdown(rows, cancel)}"})
    finally:
        workbook.close()
    if not pages:
        raise ConfigurationError("Excel workbook contains no readable values")
    return ParsedDocument(path.name, pages, [], "local_excel")
